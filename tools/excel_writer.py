import os
import json
from pathlib import Path
from typing import List, Dict, Optional

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ExcelWriter:
    def __init__(self, output_dir: str, template_path: Optional[str] = None):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.template_path = Path(template_path) if template_path else None
        
        self.headers = [
            "序号", "文件类型", "文件名称", "项目名称", "投资时间", 
            "签约时间", "决议内容", "表决结果", "合同编号", 
            "主要内容摘要", "关联合同", "备注"
        ]

    def create_headers(self, ws):
        for col, header in enumerate(self.headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border

    def set_column_widths(self, ws):
        column_widths = [8, 12, 25, 20, 15, 15, 40, 15, 15, 40, 15, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    def write_data(self, ws, data: List[Dict]):
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_idx, item in enumerate(data, 2):
            row_data = [
                item.get("序号", row_idx - 1),
                item.get("文件类型", ""),
                item.get("文件名称", ""),
                item.get("项目名称", ""),
                item.get("投资时间", ""),
                item.get("签约时间", ""),
                item.get("决议内容", ""),
                item.get("表决结果", ""),
                item.get("合同编号", ""),
                item.get("主要内容摘要", ""),
                item.get("关联合同", ""),
                item.get("备注", "")
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = thin_border

    def write_project_table(self, data: List[Dict], filename: str = "项目表.xlsx") -> str:
        if not EXCEL_AVAILABLE:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")
        
        output_path = self.output_dir / filename
        
        if self.template_path and self.template_path.exists():
            wb = load_workbook(self.template_path)
            ws = wb.active
            ws.delete_rows(2, ws.max_row)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "尽调项目表"
            self.create_headers(ws)
        
        self.set_column_widths(ws)
        self.write_data(ws, data)
        
        wb.save(output_path)
        return str(output_path)

    def merge_tables(self, existing_path: str, new_data: List[Dict], filename: str = "项目表.xlsx") -> str:
        if not EXCEL_AVAILABLE:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")
        
        existing_file = Path(existing_path)
        if not existing_file.exists():
            return self.write_project_table(new_data, filename)
        
        wb = load_workbook(existing_file)
        ws = wb.active
        
        existing_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                existing_data.append({
                    "序号": row[0] if row[0] else "",
                    "文件类型": row[1] if row[1] else "",
                    "文件名称": row[2] if row[2] else "",
                    "项目名称": row[3] if row[3] else "",
                    "投资时间": row[4] if row[4] else "",
                    "签约时间": row[5] if row[5] else "",
                    "决议内容": row[6] if row[6] else "",
                    "表决结果": row[7] if row[7] else "",
                    "合同编号": row[8] if row[8] else "",
                    "主要内容摘要": row[9] if row[9] else "",
                    "关联合同": row[10] if row[10] else "",
                    "备注": row[11] if row[11] else ""
                })
        
        merged_data = existing_data + new_data
        for idx, item in enumerate(merged_data, 1):
            item["序号"] = idx
        
        return self.write_project_table(merged_data, filename)


if __name__ == "__main__":
    import sys
    
    output_dir = sys.argv[1] if len(sys.argv) > 1 else "output"
    data_json = sys.argv[2] if len(sys.argv) > 2 else None
    template_path = sys.argv[3] if len(sys.argv) > 3 else None
    merge_path = sys.argv[4] if len(sys.argv) > 4 else None
    
    writer = ExcelWriter(output_dir, template_path)
    
    if data_json:
        data = json.loads(data_json)
        if merge_path:
            result = writer.merge_tables(merge_path, data)
        else:
            result = writer.write_project_table(data)
        print(result)
